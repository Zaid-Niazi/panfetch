from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

options = webdriver.ChromeOptions()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(options=options)

# Load the Excel file
workbook = load_workbook("D:\Work\Experiments\pan.xlsx")
sheet = workbook.active

# Go to the website
driver.get("https://www.tdscpc.gov.in/app/ded/panverify.xhtml")

# Iterate through the rows in the Excel file
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    pan_number = row[0]
    if not pan_number:
        break
    # Enter the PAN number in the input field
    pan_input = driver.find_element(By.ID, "pannumber")
    pan_input.clear()
    pan_input.send_keys(pan_number)

    # Select the correct option from the dropdown menu
    dropdown = driver.find_element(By.ID, "frmType1")
    options = dropdown.find_elements(By.TAG_NAME, "option")
    for option in options:
        if option.get_attribute("value") == "24Q":
            option.click()
            break

    # Trigger the search
    driver.find_element(By.ID, "clickGo1").click()
    # Get the text value of the "status" and "name" elements
    status = driver.find_element(By.ID, "status").text
    name = driver.find_element(By.ID, "name").text
    # Set the value of the corresponding cells in the Excel file
    sheet.cell(row=i+1, column=2).value = status
    sheet.cell(row=i+1, column=3).value = name
    # Save the excel file
    workbook.save("D:\Work\Experiments\pan.xlsx")
